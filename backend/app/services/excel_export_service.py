"""Excel export service for security review reports."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Severity colors
SEVERITY_COLORS = {
    "critical": "FF0000",  # Red
    "high": "FF8C00",  # Orange
    "medium": "FFD700",  # Yellow/Gold
    "low": "32CD32",  # Green
}

SEVERITY_BG_COLORS = {
    "critical": "FFCCCC",  # Light red
    "high": "FFE4CC",  # Light orange
    "medium": "FFFFCC",  # Light yellow
    "low": "CCFFCC",  # Light green
}

# Style constants
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")  # Blue
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def format_duration(ms: int) -> str:
    """Format milliseconds to human-readable duration."""
    if ms is None:
        return "-"
    seconds = ms // 1000
    minutes = seconds // 60
    remaining_seconds = seconds % 60
    if minutes > 0:
        return f"{minutes}分{remaining_seconds}秒"
    return f"{remaining_seconds}秒"


def format_datetime(dt: datetime) -> str:
    """Format datetime to Japanese format."""
    if dt is None:
        return "-"
    return dt.strftime("%Y年%m月%d日 %H:%M:%S")


def auto_adjust_column_width(ws, min_width: int = 8, max_width: int = 60):
    """Auto-adjust column widths based on content."""
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            if cell.value:
                cell_length = len(str(cell.value))
                # Account for Japanese characters (wider)
                if any(ord(c) > 127 for c in str(cell.value)):
                    cell_length = int(cell_length * 1.5)
                max_length = max(max_length, cell_length)

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def create_summary_sheet(wb: Workbook, review_data: dict) -> None:
    """Create the summary sheet."""
    ws = wb.active
    ws.title = "サマリー"

    # Title
    ws["A1"] = "セキュリティレビュー レポート"
    ws["A1"].font = Font(bold=True, size=18)
    ws.merge_cells("A1:E1")

    # Basic info section
    ws["A3"] = "基本情報"
    ws["A3"].font = TITLE_FONT

    info_rows = [
        ("リポジトリ", review_data.get("repo_name", "-")),
        ("ブランチ", review_data.get("branch", "main")),
        ("実行日時", review_data.get("executed_at", "-")),
        ("実行時間", review_data.get("duration", "-")),
    ]

    for i, (label, value) in enumerate(info_rows, start=4):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = value

    # Overall score section
    ws["A9"] = "総合スコア"
    ws["A9"].font = TITLE_FONT

    score = review_data.get("overall_score", 0)
    ws["A10"] = score
    ws["A10"].font = Font(bold=True, size=48)
    ws["A10"].alignment = Alignment(horizontal="center")

    # Score color based on value
    if score >= 80:
        ws["A10"].font = Font(bold=True, size=48, color="32CD32")  # Green
    elif score >= 60:
        ws["A10"].font = Font(bold=True, size=48, color="FFD700")  # Yellow
    else:
        ws["A10"].font = Font(bold=True, size=48, color="FF0000")  # Red

    ws.merge_cells("A10:B11")

    # Severity counts section
    ws["A14"] = "重要度別件数"
    ws["A14"].font = TITLE_FONT

    severity_headers = ["重要度", "件数"]
    for col, header in enumerate(severity_headers, start=1):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    severity_data = [
        ("Critical", review_data.get("critical_count", 0), "critical"),
        ("High", review_data.get("high_count", 0), "high"),
        ("Medium", review_data.get("medium_count", 0), "medium"),
        ("Low", review_data.get("low_count", 0), "low"),
    ]

    for i, (label, count, severity_key) in enumerate(severity_data, start=16):
        cell_label = ws.cell(row=i, column=1, value=label)
        cell_label.fill = PatternFill(
            start_color=SEVERITY_BG_COLORS[severity_key],
            end_color=SEVERITY_BG_COLORS[severity_key],
            fill_type="solid",
        )
        cell_label.border = THIN_BORDER

        cell_count = ws.cell(row=i, column=2, value=count)
        cell_count.border = THIN_BORDER
        cell_count.alignment = Alignment(horizontal="center")

    # Perspective scores section
    ws["A22"] = "観点別評価"
    ws["A22"].font = TITLE_FONT

    perspective_headers = ["観点", "スコア", "評価"]
    for col, header in enumerate(perspective_headers, start=1):
        cell = ws.cell(row=23, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    perspectives = review_data.get("perspective_scores", [])
    for i, perspective in enumerate(perspectives, start=24):
        name = perspective.get("name", "-")
        score = perspective.get("score", 0)
        rating = perspective.get("rating", "-")

        ws.cell(row=i, column=1, value=name).border = THIN_BORDER
        ws.cell(row=i, column=2, value=f"{score}%").border = THIN_BORDER
        ws.cell(row=i, column=3, value=rating).border = THIN_BORDER

    auto_adjust_column_width(ws)


def create_findings_list_sheet(wb: Workbook, findings: list[dict]) -> None:
    """Create the findings list sheet."""
    ws = wb.create_sheet("指摘事項一覧")

    # Headers
    headers = [
        "ID",
        "重要度",
        "タイトル",
        "ファイルパス",
        "行番号",
        "ASVS要件",
        "CWE ID",
        "検出元エージェント",
        "対応状況",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, finding in enumerate(findings, start=2):
        # ID
        cell_id = ws.cell(row=row_idx, column=1, value=finding.get("id", f"F{row_idx - 1}"))
        cell_id.border = THIN_BORDER

        # Severity with color
        severity = finding.get("severity", "medium").lower()
        severity_display = severity.capitalize()
        cell_severity = ws.cell(row=row_idx, column=2, value=severity_display)
        if severity in SEVERITY_BG_COLORS:
            cell_severity.fill = PatternFill(
                start_color=SEVERITY_BG_COLORS[severity],
                end_color=SEVERITY_BG_COLORS[severity],
                fill_type="solid",
            )
        cell_severity.border = THIN_BORDER
        cell_severity.alignment = Alignment(horizontal="center")

        # Title
        cell_title = ws.cell(row=row_idx, column=3, value=finding.get("title", "-"))
        cell_title.border = THIN_BORDER

        # File path
        location = finding.get("location", {})
        file_path = location.get("file", "-") if isinstance(location, dict) else "-"
        cell_file = ws.cell(row=row_idx, column=4, value=file_path)
        cell_file.border = THIN_BORDER

        # Line number
        line_num = location.get("line", "-") if isinstance(location, dict) else "-"
        cell_line = ws.cell(row=row_idx, column=5, value=line_num)
        cell_line.border = THIN_BORDER
        cell_line.alignment = Alignment(horizontal="center")

        # ASVS requirement
        asvs = finding.get("asvs_id", "-")
        cell_asvs = ws.cell(row=row_idx, column=6, value=asvs)
        cell_asvs.border = THIN_BORDER

        # CWE ID
        cwe = finding.get("cwe_id", "-")
        cell_cwe = ws.cell(row=row_idx, column=7, value=cwe)
        cell_cwe.border = THIN_BORDER

        # Agent
        agent = finding.get("source_agent", "SpecComplianceAgent")
        cell_agent = ws.cell(row=row_idx, column=8, value=agent)
        cell_agent.border = THIN_BORDER

        # Status
        status = finding.get("status", "open")
        status_display = "対応済み" if status == "resolved" else "未対応"
        cell_status = ws.cell(row=row_idx, column=9, value=status_display)
        cell_status.border = THIN_BORDER
        cell_status.alignment = Alignment(horizontal="center")

    auto_adjust_column_width(ws)

    # Freeze header row
    ws.freeze_panes = "A2"


def create_details_sheet(wb: Workbook, findings: list[dict]) -> None:
    """Create the detailed findings sheet."""
    ws = wb.create_sheet("詳細")

    current_row = 1

    for idx, finding in enumerate(findings, start=1):
        # Finding header
        ws.cell(row=current_row, column=1, value=f"指摘事項 #{idx}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        current_row += 1

        # Title
        title = finding.get("title", "-")
        ws.cell(row=current_row, column=1, value="タイトル")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=title)
        ws.merge_cells(f"B{current_row}:F{current_row}")
        current_row += 1

        # Severity
        severity = finding.get("severity", "medium").capitalize()
        ws.cell(row=current_row, column=1, value="重要度")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        cell_severity = ws.cell(row=current_row, column=2, value=severity)
        severity_key = severity.lower()
        if severity_key in SEVERITY_BG_COLORS:
            cell_severity.fill = PatternFill(
                start_color=SEVERITY_BG_COLORS[severity_key],
                end_color=SEVERITY_BG_COLORS[severity_key],
                fill_type="solid",
            )
        current_row += 1

        # Location
        location = finding.get("location", {})
        if isinstance(location, dict):
            loc_str = f"{location.get('file', '-')}:{location.get('line', '-')}"
        else:
            loc_str = "-"
        ws.cell(row=current_row, column=1, value="場所")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=loc_str)
        current_row += 1

        # ASVS/CWE
        asvs = finding.get("asvs_id", "-")
        cwe = finding.get("cwe_id", "-")
        ws.cell(row=current_row, column=1, value="ASVS")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=2, value=asvs)
        ws.cell(row=current_row, column=3, value="CWE")
        ws.cell(row=current_row, column=3).font = Font(bold=True)
        ws.cell(row=current_row, column=4, value=cwe)
        current_row += 1

        # Explanation
        explanation = finding.get("explanation", finding.get("description", "-"))
        ws.cell(row=current_row, column=1, value="詳細説明")
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        ws.cell(row=current_row, column=1, value=explanation)
        ws.merge_cells(f"A{current_row}:F{current_row}")
        ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
        current_row += 1

        # Vulnerable code
        code = finding.get("vulnerable_code", finding.get("code_snippet", ""))
        if code:
            ws.cell(row=current_row, column=1, value="検出されたコード")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            code_cell = ws.cell(row=current_row, column=1, value=code)
            code_cell.font = Font(name="Consolas", size=9)
            code_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            current_row += 1

        # AI Explanation
        ai_explanation = finding.get("ai_explanation", "")
        if ai_explanation:
            ws.cell(row=current_row, column=1, value="AIによる解説")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            ws.cell(row=current_row, column=1, value=ai_explanation)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            current_row += 1

        # Remediation
        remediation = finding.get("remediation", finding.get("fix_suggestion", ""))
        if remediation:
            ws.cell(row=current_row, column=1, value="修正案")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            rem_cell = ws.cell(row=current_row, column=1, value=remediation)
            rem_cell.font = Font(name="Consolas", size=9)
            rem_cell.alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{current_row}:F{current_row}")
            current_row += 1

        # References
        references = finding.get("references", [])
        if references:
            ws.cell(row=current_row, column=1, value="参考リンク")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1

            for ref in references:
                if isinstance(ref, dict):
                    ref_text = f"{ref.get('title', '')}: {ref.get('url', '')}"
                else:
                    ref_text = str(ref)
                ws.cell(row=current_row, column=1, value=ref_text)
                ws.merge_cells(f"A{current_row}:F{current_row}")
                current_row += 1

        # Separator
        current_row += 2

    auto_adjust_column_width(ws)


def create_excel_report(review_data: dict, findings: list[dict]) -> bytes:
    """Create Excel report and return as bytes.

    Args:
        review_data: Review summary data
        findings: List of finding dictionaries

    Returns:
        Excel file as bytes
    """
    wb = Workbook()

    # Create sheets
    create_summary_sheet(wb, review_data)
    create_findings_list_sheet(wb, findings)
    create_details_sheet(wb, findings)

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
